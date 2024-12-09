import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from datetime import datetime
import os
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

# Funciones
def crear_o_cargar_excel(ruta_archivo):
    if not os.path.exists(ruta_archivo):
        wb = openpyxl.Workbook()
        wb.save(ruta_archivo)
    wb = openpyxl.load_workbook(ruta_archivo)
    mes_actual = datetime.now().strftime("%B_%Y")
    if mes_actual not in wb.sheetnames:
        ws = wb.create_sheet(title=mes_actual)
        encabezados = ["Marca", "Modelo", "Precio", "RUT Cliente", "Fecha y Hora", "Estado"]
        for col_num, encabezado in enumerate(encabezados, start=1):
            col_letra = get_column_letter(col_num)
            ws[f"{col_letra}1"] = encabezado
            ws[f"{col_letra}1"].alignment = Alignment(horizontal="center", vertical="center")
    return wb

def obtener_hoja_mes_actual(wb):
    mes_actual = datetime.now().strftime("%B_%Y")
    if mes_actual not in wb.sheetnames:
        ws = wb.create_sheet(title=mes_actual)
        encabezados = ["Marca", "Modelo", "Precio", "RUT Cliente", "Fecha y Hora", "Estado"]
        for col_num, encabezado in enumerate(encabezados, start=1):
            col_letra = get_column_letter(col_num)
            ws[f"{col_letra}1"] = encabezado
            ws[f"{col_letra}1"].alignment = Alignment(horizontal="center", vertical="center")
    else:
        ws = wb[mes_actual]
    return ws


def registrar_auto(ruta_archivo, marca, modelo, precio, rut_cliente, estado):
    if not marca or not modelo or not precio or not rut_cliente or not estado:
        messagebox.showerror("Error", "Todos los campos son obligatorios.")
        return

    if not es_precio_valido(precio):
        messagebox.showerror("Error", "El precio debe ser un número válido.")
        return

    try:
        wb = openpyxl.load_workbook(ruta_archivo)
        ws = obtener_hoja_mes_actual(wb)
        proxima_fila = ws.max_row + 1
        fecha_actual = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws[f"A{proxima_fila}"] = marca
        ws[f"B{proxima_fila}"] = modelo
        ws[f"C{proxima_fila}"] = precio
        ws[f"D{proxima_fila}"] = rut_cliente
        ws[f"E{proxima_fila}"] = fecha_actual
        ws[f"F{proxima_fila}"] = estado
        wb.save(ruta_archivo)
        limpiar_formulario()
        wb = openpyxl.load_workbook(ruta_archivo)  # Recargamos el archivo
        cargar_hojas(ruta_archivo, wb)  # Actualizamos las hojas
        messagebox.showinfo("Éxito", "Registro guardado exitosamente.")
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo guardar el registro: {e}")

def cargar_datos(ruta_archivo, hoja):
    try:
        wb = openpyxl.load_workbook(ruta_archivo)
        if hoja in wb.sheetnames:
            ws = wb[hoja]
            for row in tree.get_children():
                tree.delete(row)
            for row in ws.iter_rows(min_row=2, values_only=True):
                tree.insert("", tk.END, values=row)
        else:
            messagebox.showinfo("Sin datos", "No hay datos para la hoja seleccionada.")
    except Exception as e:
        messagebox.showerror("Error", f"No se pudieron cargar los datos: {e}")

def seleccionar_ruta():
    ruta = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Archivos Excel", "*.xlsx")],
        title="Seleccione la ruta del archivo"
    )
    if ruta:
        wb = crear_o_cargar_excel(ruta)  # Cambié la llamada a la función para obtener wb
        lbl_ruta.config(text=f"Archivo: {ruta}")
        btn_registrar.config(state=tk.NORMAL)
        cargar_hojas(ruta, wb)  # Pasamos wb para cargar las hojas
    return ruta

def eliminar_registro(ruta_archivo):
    try:
        selected_item = tree.selection()
        if not selected_item:
            messagebox.showerror("Error", "Seleccione un registro para eliminar.")
            return

        wb = openpyxl.load_workbook(ruta_archivo)
        hoja = combo_hojas.get()
        ws = wb[hoja]
        fila = tree.index(selected_item[0]) + 2  # +2 porque las filas de Excel comienzan en 1 y hay un encabezado

        ws.delete_rows(fila)

        wb.save(ruta_archivo)
        cargar_datos(ruta_archivo, hoja)
        messagebox.showinfo("Éxito", "Registro eliminado exitosamente.")
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo eliminar el registro: {e}")

def cargar_hojas(ruta_archivo, wb):
    try:
        hojas = wb.sheetnames
        if hojas:
            combo_hojas['values'] = hojas
            combo_hojas.current(hojas.index(datetime.now().strftime("%B_%Y")))  # Selecciona la hoja del mes actual
            cargar_datos(ruta_archivo, combo_hojas.get())
        else:
            messagebox.showinfo("Sin hojas", "No hay hojas en el archivo.")
    except Exception as e:
        messagebox.showerror("Error", f"No se pudieron cargar las hojas: {e}")

def limpiar_formulario():
    entry_marca.delete(0, tk.END)
    entry_modelo.delete(0, tk.END)
    entry_precio.delete(0, tk.END)
    entry_rut.delete(0, tk.END)
    estado_var.set("")

def es_precio_valido(precio):
    try:
        float(precio)
        return True
    except ValueError:
        return False
def seleccionar_registro(event):
    for item in tree.selection():
        valores = tree.item(item, "values")
        if valores:
            entry_marca.delete(0, tk.END)
            entry_modelo.delete(0, tk.END)
            entry_precio.delete(0, tk.END)
            entry_rut.delete(0, tk.END)
            estado_var.set(valores[5])

            entry_marca.insert(0, valores[0])
            entry_modelo.insert(0, valores[1])
            entry_precio.insert(0, valores[2])
            entry_rut.insert(0, valores[3])

def actualizar_registro(ruta_archivo):
    try:
        selected_item = tree.selection()
        if not selected_item:
            messagebox.showerror("Error", "Seleccione un registro para editar.")
            return

        marca = entry_marca.get()
        modelo = entry_modelo.get()
        precio = entry_precio.get()
        rut_cliente = entry_rut.get()
        estado = estado_var.get()

        if not marca or not modelo or not precio or not rut_cliente or not estado:
            messagebox.showerror("Error", "Todos los campos son obligatorios.")
            return

        wb = openpyxl.load_workbook(ruta_archivo)
        hoja = combo_hojas.get()
        ws = wb[hoja]
        fila = tree.index(selected_item[0]) + 2

        ws[f"A{fila}"] = marca
        ws[f"B{fila}"] = modelo
        ws[f"C{fila}"] = precio
        ws[f"D{fila}"] = rut_cliente
        ws[f"F{fila}"] = estado

        wb.save(ruta_archivo)
        limpiar_formulario()
        cargar_datos(ruta_archivo, hoja)
        messagebox.showinfo("Éxito", "Registro actualizado exitosamente.")
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo actualizar el registro: {e}")

# Interfaz gráfica
root = tk.Tk()
root.title("Registro de Autos")

lbl_ruta = tk.Label(root, text="Seleccione el archivo para guardar los registros", wraplength=400)
lbl_ruta.pack(pady=10)

btn_seleccionar = tk.Button(root, text="Seleccionar Archivo", command=seleccionar_ruta)
btn_seleccionar.pack(pady=5)

frame_form = tk.Frame(root)
frame_form.pack(pady=10)

tk.Label(frame_form, text="Marca:").grid(row=0, column=0, padx=5, pady=5)
entry_marca = tk.Entry(frame_form)
entry_marca.grid(row=0, column=1, padx=5, pady=5)

tk.Label(frame_form, text="Modelo:").grid(row=1, column=0, padx=5, pady=5)
entry_modelo = tk.Entry(frame_form)
entry_modelo.grid(row=1, column=1, padx=5, pady=5)

tk.Label(frame_form, text="Precio:").grid(row=2, column=0, padx=5, pady=5)
entry_precio = tk.Entry(frame_form)
entry_precio.grid(row=2, column=1, padx=5, pady=5)

tk.Label(frame_form, text="RUT Cliente:").grid(row=3, column=0, padx=5, pady=5)
entry_rut = tk.Entry(frame_form)
entry_rut.grid(row=3, column=1, padx=5, pady=5)

tk.Label(frame_form, text="Estado:").grid(row=4, column=0, padx=5, pady=5)
estado_var = tk.StringVar()
combo_estado = ttk.Combobox(frame_form, textvariable=estado_var, values=["Retirado", "Sin retirar"], state="readonly")
combo_estado.grid(row=4, column=1, padx=5, pady=5)

btn_registrar = tk.Button(root, text="Registrar Auto", state=tk.DISABLED, command=lambda: registrar_auto(
    lbl_ruta.cget("text").replace("Archivo: ", ""),
    entry_marca.get(),
    entry_modelo.get(),
    entry_precio.get(),
    entry_rut.get(),
    estado_var.get()
))
btn_registrar.pack(pady=5)

btn_actualizar = tk.Button(root, text="Actualizar Registro", state=tk.NORMAL, command=lambda: actualizar_registro(
    lbl_ruta.cget("text").replace("Archivo: ", "")
))
btn_actualizar.pack(pady=5)

btn_eliminar = tk.Button(root, text="Eliminar Registro", state=tk.NORMAL, command=lambda: eliminar_registro(
    lbl_ruta.cget("text").replace("Archivo: ", "")
))
btn_eliminar.pack(pady=5)
# Desplegable para seleccionar hoja
combo_hojas = ttk.Combobox(root, state="readonly")
combo_hojas.pack(pady=10)
combo_hojas.bind("<<ComboboxSelected>>", lambda event: cargar_datos(lbl_ruta.cget("text").replace("Archivo: ", ""), combo_hojas.get()))

tree = ttk.Treeview(root, columns=("Marca", "Modelo", "Precio", "RUT Cliente", "Fecha y Hora", "Estado"), show="headings")
tree.pack(pady=10, fill=tk.BOTH, expand=True)


for col in tree["columns"]:
    tree.heading(col, text=col)
    tree.column(col, anchor=tk.W)

tree.bind("<Double-1>", seleccionar_registro)

firma = tk.Label(root, text="Power By RIAG", font=("Arial", 8), fg="gray")
firma.place(relx=1.0, rely=1.0, anchor="se")  # Posiciona en la esquina inferior derecha

#By Rafael Ingacio Aburto Garrido
root.mainloop()


